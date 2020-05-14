import openpyxl


class HomePageData:

    test_homepae_data =[{"firstname":"Dipty","email":"Dash","gender":"Female"},{"firstname":"Dispty","email":"Dsash","gender":"Male"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\sasneha\\Desktop\\Python Demo.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name  :
                for j in range(2, sheet.max_column + 1):
                    # print(sheet.cell(row=i, column=j).value)# we will store this value in dictionary
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]
