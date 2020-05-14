import openpyxl
book = openpyxl.load_workbook("C:\\Users\\sasneha\\Desktop\\Python Demo.xlsx")
sheet = book.active
Dict ={}
cell = sheet.cell(row=2,column=3)
print(cell.value)
sheet.cell(row=2,column=3).value ="Dipty"

print(sheet.cell(row=2,column=3).value)

print(sheet.max_column)
print(sheet.max_row)

print(sheet['A7'].value)

for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value=="Testcase2":
        for j in range(1, sheet.max_column + 1):
            #print(sheet.cell(row=i, column=j).value)# we will store this value in dictionary
            Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value
print(Dict)



